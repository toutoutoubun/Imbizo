"""Spreadsheet transcript importer for XLSX and ODS."""

from __future__ import annotations

import uuid
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Iterable

from imbizo.app.errors import ImportFailure
from imbizo.domain.transcripts import SegmentLevel, SourceFormat, TranscriptDocument, TranscriptSegment, split_tokens_preserving_offsets
from imbizo.importers.base import ImportedBundle, ImportOptions

TEXT_KEYS = (
    "text",
    "transcript",
    "transcription",
    "utterance",
    "utterance_text",
    "segment_text",
    "utterance_segment_transcription",
)
TOKEN_KEYS = ("token", "tokens", "token_text", "word", "surface", "form", "語", "単語", "トークン")
LANGUAGE_KEYS = ("language", "language_code", "lang", "lang_id", "utterance_segment_lang_id")
START_KEYS = ("start_ms", "start", "start_time", "begin", "begin_ms")
END_KEYS = ("end_ms", "end", "end_time", "finish", "finish_ms")
DURATION_KEYS = ("duration_ms", "duration", "utterance_segment_duration")
SPEAKER_KEYS = ("speaker", "speaker_id", "utterance_speaker_id")
ID_KEYS = ("utterance_id", "segment_id", "id")
HEADER_SCAN_NON_EMPTY_ROWS = 50


def _to_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _normalise_header_part(value: str) -> str:
    return "".join(character.lower() if character.isalnum() else "_" for character in value.replace("#", "")).strip("_")


def _header_keys(value: object) -> list[str]:
    raw = str(value or "").strip()
    if not raw:
        return []
    parts = [_normalise_header_part(part) for part in raw.strip("/").split("/") if _normalise_header_part(part)]
    keys: list[str] = []
    if parts:
        keys.append("_".join(parts))
        keys.append(parts[-1])
        if len(parts) >= 2:
            keys.append(f"{parts[-2]}_{parts[-1]}")
        if len(parts) >= 3:
            keys.append("_".join(parts[-3:]))
    simple = _normalise_header_part(raw)
    if simple:
        keys.append(simple)
    unique: list[str] = []
    for key in keys:
        if key and key not in unique:
            unique.append(key)
    return unique


def _looks_like_header(row: Iterable[object]) -> bool:
    keys = {key for value in row for key in _header_keys(value)}
    return any(key in keys for key in (*TEXT_KEYS, *TOKEN_KEYS))


def _first_non_empty_cell(row: Iterable[object]) -> object | None:
    """Return the first non-empty cell from a row without assuming headers."""

    for value in row:
        if value not in (None, ""):
            return value
    return None


def _rows_from_matrix(matrix: list[tuple[object, ...] | list[object]]) -> tuple[list[dict[str, object]], int]:
    if not matrix:
        return [], 0
    detected_header_index = None
    non_empty_seen = 0
    for index, row in enumerate(matrix):
        if _first_non_empty_cell(row) in (None, ""):
            continue
        non_empty_seen += 1
        if _looks_like_header(row):
            detected_header_index = index
            break
        if non_empty_seen >= HEADER_SCAN_NON_EMPTY_ROWS:
            break
    if detected_header_index is None:
        parsed_no_header_rows = [
            {"text": value}
            for row in matrix
            if (value := _first_non_empty_cell(row)) not in (None, "")
        ]
        return parsed_no_header_rows, 0
    header_index = detected_header_index
    headers = list(matrix[header_index])
    parsed_rows: list[dict[str, object]] = []
    for row in matrix[header_index + 1 :]:
        mapped: dict[str, object] = {}
        for header, value in zip(headers, row, strict=False):
            for key in _header_keys(header):
                mapped.setdefault(key, value)
        if any(value not in (None, "") for value in mapped.values()):
            parsed_rows.append(mapped)
    return parsed_rows, header_index + 1


def _first(row: dict[str, object], keys: tuple[str, ...]) -> object:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return None


def _language_code(value: object) -> str:
    raw = str(value or "").strip().lower()
    return {
        "af": "afr",
        "afrikaans": "afr",
        "en": "eng",
        "english": "eng",
        "sesotho": "sot",
        "st": "sot",
        "sotho": "sot",
        "tn": "tsn",
        "setswana": "tsn",
        "xh": "xho",
        "isixhosa": "xho",
        "xhosa": "xho",
        "zu": "zul",
        "isizulu": "zul",
        "zulu": "zul",
    }.get(raw, raw)


class SpreadsheetImporter:
    """Import XLSX and ODS spreadsheet transcripts."""

    name = "spreadsheet"

    def can_import(self, path: Path) -> bool:
        """Return whether this importer can parse a copied local file."""

        return path.suffix.lower() in {".xlsx", ".ods"}

    def import_file(self, path: Path, options: ImportOptions) -> ImportedBundle:
        """Parse spreadsheet rows into segments and tokens."""

        rows = self._read_xlsx(path) if path.suffix.lower() == ".xlsx" else self._read_ods(path)
        document = TranscriptDocument(
            id=str(uuid.uuid4()),
            name=path.stem,
            source_format=SourceFormat.XLSX if path.suffix.lower() == ".xlsx" else SourceFormat.ODS,
            media_asset_id=options.linked_media_asset_id,
            relative_path=str(path),
            original_filename=path.name,
        )
        segments: list[TranscriptSegment] = []
        tokens = []
        token_language_codes: dict[str, str] = {}
        text_sources: set[str] = set()
        for order, row in enumerate(rows, start=1):
            text_source = "text"
            text = str(_first(row, TEXT_KEYS) or "").strip()
            if not text:
                text = str(_first(row, TOKEN_KEYS) or "").strip()
                text_source = "token"
            if not text:
                continue
            text_sources.add(text_source)
            start_ms = _to_int(_first(row, START_KEYS))
            end_ms = _to_int(_first(row, END_KEYS))
            duration_ms = _to_int(_first(row, DURATION_KEYS))
            if end_ms is None and start_ms is not None and duration_ms is not None:
                end_ms = start_ms + duration_ms
            elif end_ms is None and start_ms is None and duration_ms is not None:
                start_ms = 0
                end_ms = duration_ms
            speaker_value = str(_first(row, SPEAKER_KEYS) or "").strip() or None
            segment = TranscriptSegment(
                id=str(uuid.uuid4()),
                transcript_document_id=document.id,
                media_asset_id=options.linked_media_asset_id,
                segment_level=SegmentLevel.UTTERANCE,
                sort_order=order,
                text_original=text,
                start_ms=start_ms,
                end_ms=end_ms,
                speaker_id=speaker_value,
                external_ref=str(_first(row, ID_KEYS) or ""),
            )
            segments.append(segment)
            segment_tokens = split_tokens_preserving_offsets(segment.id, text)
            language_code = _language_code(_first(row, LANGUAGE_KEYS))
            if language_code:
                for token in segment_tokens:
                    token_language_codes[token.id] = language_code
            tokens.extend(segment_tokens)
        return ImportedBundle(
            document=document,
            segments=segments,
            tokens=tokens,
            token_language_codes=token_language_codes,
            report={
                "segments": len(segments),
                "tokens": len(tokens),
                "imported_language_labels": len(token_language_codes),
                "text_source": ",".join(sorted(text_sources)) if text_sources else None,
            },
        )

    def _read_xlsx(self, path: Path) -> list[dict[str, object]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportFailure("XLSX import requires the optional openpyxl package.") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            parsed, _header_row = _rows_from_matrix(rows)
            return parsed
        finally:
            workbook.close()

    def _read_ods(self, path: Path) -> list[dict[str, object]]:
        with zipfile.ZipFile(path) as archive:
            content = archive.read("content.xml")
        root = ET.fromstring(content)
        ns = {
            "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
            "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
        }
        table = root.find(".//table:table", ns)
        if table is None:
            return []
        matrix: list[list[str]] = []
        for row in table.findall("table:table-row", ns):
            values: list[str] = []
            for cell in row.findall("table:table-cell", ns):
                repeat = int(cell.attrib.get("{urn:oasis:names:tc:opendocument:xmlns:table:1.0}number-columns-repeated", "1"))
                text_parts = [node.text or "" for node in cell.findall(".//text:p", ns)]
                value = "\n".join(text_parts)
                values.extend([value] * repeat)
            if any(value for value in values):
                matrix.append(values)
        if not matrix:
            return []
        parsed, _header_row = _rows_from_matrix(matrix)
        return parsed
